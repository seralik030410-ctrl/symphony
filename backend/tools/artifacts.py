from __future__ import annotations

import asyncio
import csv
import io
from pydantic import Field

from backend.tools.contracts import Tool, ToolContext, ToolInput, ToolError, ToolResult
from backend.artifacts.schemas import Format, SCHEMAS, EXAMPLES
from backend.artifacts.store import ArtifactStore
from backend.artifacts.store import digest, MAX_SPEC_BYTES
from backend.storage.repository import NotFoundError


class SchemaInput(ToolInput):
    format: Format


class DocumentSchemaTool(Tool):
    name = "artifact.schema"
    title = "Read document schema"
    description = "Read the trusted document JSON schema and example for pdf/xlsx/docx/pptx. Call before fs.write of the spec, then artifact.render. No generated renderer code is needed."
    input_model = SchemaInput

    async def execute(self, context, arguments):
        return ToolResult({"schema": SCHEMAS[arguments.format].model_json_schema(), "example": EXAMPLES[arguments.format], "workflow": "Write JSON using fs.write, then artifact.render with format and spec_path. To revise, pass the returned artifact_id; each version is immutable. PDF/DOCX/PPTX have real page previews. XLSX: worksheet row 1 is the generated header, so rows[0] is Excel row 2; the Nth rows item is Excel row N+2. A formula target must be uppercase, point to a null cell in rows, and ranges start at row 2 (not header row 1). Example with data rows 2,3 and total row 4: rows [[\"A\",10],[\"B\",20],[\"Total\",null]], formulas {\"B4\":\"=SUM(B2:B3)\"}. SUM/AVERAGE/MIN/MAX/COUNT/ABS/ROUND and + - * / only. No network or external files."})


class RenderInput(ToolInput):
    format: Format
    spec_path: str = Field(min_length=1, max_length=500)
    artifact_id: str | None = Field(default=None, pattern=r"^[0-9a-f]{32}$")


class RenderDocumentTool(Tool):
    name = "artifact.render"
    title = "Render and validate document"
    description = "Create PDF, XLSX, DOCX or PPTX from a JSON spec file in this chat. Read artifact.schema first. Trusted offline renderer validates and saves source, recipe, file, previews and version. artifact_id revises an existing artifact. Return only confirmed download_url."
    input_model = RenderInput
    read_only = False
    timeout_seconds = 180

    def __init__(self, store: ArtifactStore): self.store = store

    def dependency_fingerprint(self, context, arguments):
        try:
            path = self.store.workspaces.resolve(context.session_id, str(arguments.get("spec_path", "")), must_exist=True)
            if path.is_file() and path.stat().st_size <= MAX_SPEC_BYTES: return digest(path)
        except (ToolError, OSError): pass
        return None

    async def execute(self, context: ToolContext, arguments: RenderInput):
        try: return ToolResult(await self.store.render(context, arguments.format, arguments.spec_path, arguments.artifact_id))
        except NotFoundError as error: raise ToolError("not_found", str(error)) from error


class ArtifactInspectInput(ToolInput):
    artifact_id: str | None = Field(default=None, pattern=r"^[0-9a-f]{32}$")
    version: int | None = Field(default=None, ge=1)


class InspectArtifactTool(Tool):
    name = "artifact.inspect"
    title = "Inspect saved documents"
    description = "List documents/versions in this chat, or read a document's saved JSON source and validation summary to revise it. Does not read documents from other chats."
    input_model = ArtifactInspectInput

    def __init__(self, store: ArtifactStore): self.store = store

    async def execute(self, context, arguments):
        if not arguments.artifact_id:
            return ToolResult({"artifacts": self.store.list(context.session_id)[:60]})
        try: detail = self.store.get(context.session_id, arguments.artifact_id, arguments.version)
        except NotFoundError as error: raise ToolError("not_found", str(error)) from error
        path = self.store.file(context.session_id, arguments.artifact_id, detail["version"], "source.json")
        text = path.read_text(encoding="utf-8")
        return ToolResult({**self.store._compact(detail), "source": text[:12000], "source_truncated": len(text) > 12000,
            "warnings": detail["validation"].get("warnings", []), "calculation": detail["validation"].get("calculation", {})})


class ReadTableInput(ToolInput):
    path: str = Field(min_length=1, max_length=500)
    sheet: str | None = Field(default=None, max_length=31)
    offset: int = Field(default=0, ge=0, le=100_000)
    limit: int = Field(default=30, ge=1, le=100)


def read_table(workspaces, session_id, arguments):
    path = workspaces.resolve(session_id, arguments.path, must_exist=True)
    if not path.is_file() or path.stat().st_size > 8_000_000: raise ToolError("invalid_table", "Table input must be at most 8 MB")
    if path.suffix.lower() == ".csv":
        text = path.read_text(encoding="utf-8-sig")
        try: dialect = csv.Sniffer().sniff(text[:8192], delimiters=",;\t")
        except csv.Error: dialect = csv.excel
        rows, count = [], 0
        for index, row in enumerate(csv.reader(io.StringIO(text), dialect)):
            count = index + 1
            if count > 100_000: raise ToolError("invalid_table", "CSV exceeds 100,000 rows")
            if arguments.offset <= index < arguments.offset + arguments.limit: rows.append(row[:24])
        return {"rows": rows, "offset": arguments.offset, "total_rows": count, "sheet_names": [], "cached_values": False, "column_limit": 24}
    if path.suffix.lower() != ".xlsx": raise ToolError("invalid_table", "Use .csv or .xlsx (macros are not accepted)")
    from zipfile import ZipFile
    from openpyxl import load_workbook
    with ZipFile(path) as archive:
        if len(archive.infolist()) > 2000 or sum(item.file_size for item in archive.infolist()) > 30_000_000:
            raise ToolError("invalid_table", "Workbook decompressed size exceeds 30 MB")
    book = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        if arguments.sheet and arguments.sheet not in book.sheetnames: raise ToolError("invalid_sheet", "Sheet not found")
        sheet = book[arguments.sheet] if arguments.sheet else book.active
        rows = [[value.isoformat() if hasattr(value, "isoformat") else value for value in row] for row in sheet.iter_rows(min_row=arguments.offset + 1, max_row=arguments.offset + arguments.limit, max_col=min(sheet.max_column or 24, 24), values_only=True)]
        return {"sheet_names": book.sheetnames, "sheet": sheet.title, "rows": rows, "offset": arguments.offset, "total_rows": sheet.max_row, "cached_values": True, "warning": "Formula cells use saved Excel values, which can be stale or absent; no recalculation of uploaded files."}
    finally: book.close()


class ReadTableTool(Tool):
    name = "artifact.read_table"
    title = "Read table data"
    description = "Read a bounded page of CSV/XLSX data from this chat's workspace, including user uploads under inputs/. Excel values are cached, not recalculated. Use returned rows as evidence for a report; never invent missing values."
    input_model = ReadTableInput
    def __init__(self, workspaces): self.workspaces = workspaces
    async def execute(self, context, arguments):
        try: return ToolResult(await asyncio.to_thread(read_table, self.workspaces, context.session_id, arguments))
        except ToolError: raise
        except Exception as error: raise ToolError("table_read_failed", str(error)[:500]) from error
