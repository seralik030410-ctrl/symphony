#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
audit_workbook.py — read-only technical audit of a FinControl-style workbook.

НИЧЕГО не пишет и не пересчитывает файл (см. references/known-gotchas.md —
пересчёт этой книги через LibreOffice/recalc.py ломает FILTER-справочники).
Это Phase 1 (Workbook health) + часть Phase 13 (Controls) из
references/audit-checklist.md — быстрый первый проход перед ручным аудитом.

Запуск:
    python3 audit_workbook.py path/to/FinControl_Master_v1.xlsx

Выводит JSON со следующими блоками:
  - sheets: список листов, видимость, размеры
  - tables: все Excel Tables (ListObjects) по листам с диапазонами
  - named_ranges: workbook-level defined names
  - formula_errors: ячейки с кэшированным значением-ошибкой (#REF!, #NAME? и т.д.) —
    определяется БЕЗ пересчёта, по текущим сохранённым значениям
  - risky_functions: использование XLOOKUP/UNIQUE/SORT/SEQUENCE/INDIRECT/OFFSET/FILTER
    (see known-gotchas.md — не все из них "плохие", но все требуют внимания)
  - duplicate_ids: дубли в первом столбце (обычно "... ID") каждой Table
  - hardcoded_ranges: подозрительные фиксированные диапазоны вроде $A$2:$A$6 в формулах
    вне Tables (антипаттерн масштабируемости, см. §99 книги правил)
  - control_center: текущее состояние 16_Control (если лист существует) — Critical/
    Technical Critical и непустые строки проверок
  - external_links: наличие внешних ссылок на другие файлы
"""

import json
import re
import sys
import warnings

warnings.filterwarnings("ignore")

try:
    import openpyxl
    from openpyxl.utils import range_boundaries
except ImportError:
    print("Требуется openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ERROR_VALUES = {"#REF!", "#NAME?", "#VALUE!", "#DIV/0!", "#N/A", "#NULL!",
                 "#NUM!", "#CALC!", "#SPILL!", "#GETTING_DATA"}

RISKY_FUNCTIONS = {
    "XLOOKUP": "unsupported by this environment's LibreOffice recalc — never introduce",
    "XMATCH": "unsupported by this environment's LibreOffice recalc — never introduce",
    "_xlws.SORT": "spilling array fn — unsupported by this environment's recalc",
    "_xlws.UNIQUE": "spilling array fn — unsupported by this environment's recalc",
    "_xlws.FILTER": "KNOWN IN THIS WORKBOOK (dropdown source lists) — do not recalc via LibreOffice; see known-gotchas.md",
    "SEQUENCE": "spilling array fn — unsupported by this environment's recalc",
    "INDIRECT": "volatile + breaks structured-reference scalability principle",
    "OFFSET": "volatile — check if it's the known cumulative-total pattern (08_Акты) or a new one",
}

FIXED_RANGE_RE = re.compile(r"\$[A-Z]{1,3}\$\d+:\$[A-Z]{1,3}\$\d+")


def cell_formula_text(cell):
    v = cell.value
    if hasattr(v, "text"):
        return v.text
    if isinstance(v, str) and v.startswith("="):
        return v
    return None


def audit(path):
    report = {"file": path}

    wb_f = openpyxl.load_workbook(path, data_only=False, read_only=False)
    wb_v = openpyxl.load_workbook(path, data_only=True, read_only=False)

    # ---- sheets ----
    report["sheets"] = [
        {"name": s, "state": wb_f[s].sheet_state,
         "dimensions": wb_f[s].dimensions,
         "max_row": wb_f[s].max_row, "max_col": wb_f[s].max_column}
        for s in wb_f.sheetnames
    ]

    # ---- tables ----
    tables = []
    for s in wb_f.sheetnames:
        ws = wb_f[s]
        try:
            tbls = ws.tables
        except Exception:
            continue
        for tname in tbls:
            tbl = tbls[tname]
            tables.append({"sheet": s, "table": tname, "range": tbl.ref})
    report["tables"] = tables

    # ---- named ranges (workbook-level) ----
    report["named_ranges"] = [
        {"name": n, "refers_to": d.attr_text} for n, d in wb_f.defined_names.items()
    ]

    # ---- external links ----
    ext_links = []
    for s in wb_f.sheetnames:
        ws = wb_f[s]
        for row in ws.iter_rows():
            for c in row:
                t = cell_formula_text(c)
                if t and re.search(r"\[\d+\]", t):
                    ext_links.append(f"{s}!{c.coordinate}: {t}")
    report["external_links"] = ext_links

    # ---- formula errors (from cached values — NO recalculation needed) ----
    formula_errors = []
    for s in wb_v.sheetnames:
        ws = wb_v[s]
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value in ERROR_VALUES:
                    formula_errors.append({"cell": f"{s}!{c.coordinate}", "error": c.value})
    report["formula_errors"] = {
        "count": len(formula_errors),
        "note": "detected from cached values already stored in the file — do NOT "
                "run a generic recalc to check this, see known-gotchas.md",
        "cells": formula_errors[:200],
    }

    # ---- risky / notable functions ----
    risky_hits = {}
    for s in wb_f.sheetnames:
        ws = wb_f[s]
        for row in ws.iter_rows():
            for c in row:
                t = cell_formula_text(c)
                if not t:
                    continue
                for fn in RISKY_FUNCTIONS:
                    if fn in t:
                        risky_hits.setdefault(fn, []).append(f"{s}!{c.coordinate}")
    report["risky_functions"] = {
        fn: {"count": len(cells), "note": RISKY_FUNCTIONS[fn], "cells": cells[:30]}
        for fn, cells in risky_hits.items()
    }

    # ---- duplicate IDs per table ----
    # Only checked for tables with a genuine single-column primary/document key
    # (whitelisted below from the actual workbook schema). Tables whose first
    # column is a foreign key by design (Project ID in tblBvA, Counterparty ID
    # in tblAP, Cost Code repeated per line, etc.) are intentionally excluded —
    # flagging those as "duplicates" would be a false positive, see
    # references/known-gotchas.md.
    PRIMARY_KEY_COLUMNS = {
        "tblPR": "PR ID", "tblQEF": "QEF ID", "tblQEFOffers": "Offer ID",
        "tblPO": "PO ID", "tblReceipts": "Receipt ID", "tblIssue": "Issue ID",
        "tblUse": "Use ID", "tblActs": "Document ID", "tblPAYREQ": "PAYREQ ID",
        "tblCash": "Payment ID", "tblAdvances": "Advance ID",
        "tblEquip": "Equipment ID", "tblRevenue": "Revenue ID",
        "tblAR": "AR ID", "tblOtherCosts": "Cost ID",
        "tblVariations": "Variation ID", "tblAdjust": "Adjustment ID",
        "tblApprovals": "Approval ID", "tblEquipFcst": "Forecast ID",
        "tblContracts": "Contract ID", "tblContractBOQ": "Contract BOQ Link ID",
        "tblBudgetNew": "Budget Line ID", "tblBOQNew": "BOQ Code",
        "tblBOMNew": "BOM Line ID", "tblSchedule": "Activity ID",
        "tblDocs": "Document ID",
        "tblProjects": "Project ID", "tblBlocks": "Block ID",
        "tblCounterparties": "Counterparty ID", "tblMaterials": "Material ID",
        "tblUOM": "UOM Code", "tblCostCodes": "Cost Code",
        "tblCurrencies": "Currency Code", "tblEmployees": "Employee ID",
        "tblWarehouses": "Warehouse ID", "tblCashAccounts": "Account ID",
        "tblPriority": "Priority Code", "tblSettings": "Setting Name",
    }
    dup_report = []
    schema_drift = []
    for s in wb_f.sheetnames:
        ws_f = wb_f[s]
        ws_v = wb_v[s]
        try:
            tbls = ws_f.tables
        except Exception:
            continue
        for tname in tbls:
            expected_col = PRIMARY_KEY_COLUMNS.get(tname)
            if not expected_col:
                continue  # not whitelisted -> composite/FK key, skip (avoid false positives)
            tbl = tbls[tname]
            min_col, min_row, max_col, max_row = range_boundaries(tbl.ref)
            header = ws_f.cell(row=min_row, column=min_col).value
            if str(header).strip() != expected_col:
                schema_drift.append(
                    f"{s}!{tname}: expected first column '{expected_col}', found "
                    f"'{header}' — table structure changed, update PRIMARY_KEY_COLUMNS "
                    f"in this script and re-check duplicates manually for this table"
                )
                continue  # schema drifted from what this script expects — skip, don't guess
            seen = {}
            for r in range(min_row + 1, max_row + 1):
                v = ws_v.cell(row=r, column=min_col).value
                if v in (None, ""):
                    continue
                seen.setdefault(v, []).append(r)
            dups = {k: v for k, v in seen.items() if len(v) > 1}
            if dups:
                dup_report.append({
                    "sheet": s, "table": tname, "id_column": header,
                    "duplicates": {str(k): rows for k, rows in dups.items()},
                })
    report["duplicate_ids"] = dup_report
    report["schema_drift_warnings"] = schema_drift

    # ---- hardcoded fixed ranges outside of Tables (scalability smell) ----
    table_ranges = set()
    for s in wb_f.sheetnames:
        ws = wb_f[s]
        try:
            for tname in ws.tables:
                table_ranges.add((s, ws.tables[tname].ref))
        except Exception:
            pass
    hardcoded = []
    for s in wb_f.sheetnames:
        ws = wb_f[s]
        for row in ws.iter_rows():
            for c in row:
                t = cell_formula_text(c)
                if not t:
                    continue
                for m in FIXED_RANGE_RE.findall(t):
                    hardcoded.append(f"{s}!{c.coordinate}: {m}")
    report["hardcoded_ranges"] = {
        "count": len(hardcoded),
        "note": "fixed $A$2:$A$6-style ranges outside Excel Tables — check each one "
                "actually needs to be fixed-size (e.g. a small lookup table) rather "
                "than growing with new rows",
        "sample": hardcoded[:50],
    }

    # ---- Control Center snapshot, if present ----
    if "16_Control" in wb_v.sheetnames:
        ws = wb_v["16_Control"]
        cc = {"critical_controls_total": ws["B2"].value,
              "technical_critical_total": ws["B3"].value,
              "nonzero_checks": []}
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
            vals = [c.value for c in row]
            if len(vals) >= 3 and isinstance(vals[2], (int, float)) and vals[2] not in (0, None):
                cc["nonzero_checks"].append({
                    "module": vals[0], "check": vals[1], "count": vals[2],
                    "level": vals[3] if len(vals) > 3 else None,
                    "action": vals[4] if len(vals) > 4 else None,
                    "source": vals[5] if len(vals) > 5 else None,
                })
        report["control_center"] = cc

    # ---- quick verdict hint (not a substitute for the full audit checklist) ----
    verdict_hint = "PASS"
    if formula_errors or dup_report or ext_links:
        verdict_hint = "FAIL-ish: technical defects found — see formula_errors / duplicate_ids / external_links"
    elif report.get("control_center", {}).get("nonzero_checks"):
        verdict_hint = "Technical layer looks clean; Business Critical items are open — see control_center.nonzero_checks (this may be correct system behavior, not a defect — see principles.md #8-9)"
    report["quick_verdict_hint"] = verdict_hint

    return report


def main():
    if len(sys.argv) != 2:
        print("Использование: python3 audit_workbook.py path/to/workbook.xlsx", file=sys.stderr)
        sys.exit(1)
    result = audit(sys.argv[1])
    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
